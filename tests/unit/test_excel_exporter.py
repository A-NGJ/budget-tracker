from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from budget_tracker.analytics.models import (
    AnalyticsPeriod,
    AnalyticsResult,
    CategoryRow,
    MonthRow,
    SourceRow,
    SubcategoryRow,
    SummaryData,
)
from budget_tracker.config.settings import get_settings
from budget_tracker.exporters import ExcelExporter
from budget_tracker.models.transaction import StandardTransaction

SHEET_NAMES = [
    "Transactions",
    "Summary",
    "Category Breakdown",
    "Monthly Trends",
    "Source Analysis",
]


@pytest.fixture()
def sample_transactions() -> list[StandardTransaction]:
    return [
        StandardTransaction(
            date=date(2025, 10, 5),
            category="Food & Drinks",
            subcategory="Restaurants",
            amount=Decimal("-125.50"),
            source="Danske Bank",
            description="Cafe Central",
        ),
        StandardTransaction(
            date=date(2025, 10, 10),
            category="Transportation",
            subcategory="Public Transport",
            amount=Decimal("-24.00"),
            source="Nordea",
            description="Metro ticket",
        ),
        StandardTransaction(
            date=date(2025, 10, 15),
            category="Income",
            subcategory="Salary",
            amount=Decimal("30000.00"),
            source="Danske Bank",
            description="Monthly salary",
        ),
        StandardTransaction(
            date=date(2025, 11, 3),
            category="Food & Drinks",
            subcategory="Groceries",
            amount=Decimal("-200.00"),
            source="Nordea",
            description="Supermarket",
        ),
        StandardTransaction(
            date=date(2025, 11, 12),
            category="Car",
            subcategory="Fuel",
            amount=Decimal("-350.00"),
            source="Danske Bank",
            description="Gas station",
        ),
    ]


@pytest.fixture()
def sample_analytics() -> AnalyticsResult:
    period = AnalyticsPeriod(from_date=None, to_date=None, label="All Time")
    return AnalyticsResult(
        summary=SummaryData(
            total_transactions=5,
            total_income=Decimal("30000.00"),
            total_expenses=Decimal("-699.50"),
            net=Decimal("29300.50"),
            avg_transaction=Decimal("-174.88"),
            period=period,
        ),
        category_data=[
            CategoryRow(
                category="Car",
                total=Decimal("-350.00"),
                percentage=50.0,
                transaction_count=1,
                subcategories=[
                    SubcategoryRow(
                        subcategory="Fuel",
                        total=Decimal("-350.00"),
                        transaction_count=1,
                    ),
                ],
            ),
            CategoryRow(
                category="Food & Drinks",
                total=Decimal("-325.50"),
                percentage=46.5,
                transaction_count=2,
                subcategories=[
                    SubcategoryRow(
                        subcategory="Groceries",
                        total=Decimal("-200.00"),
                        transaction_count=1,
                    ),
                    SubcategoryRow(
                        subcategory="Restaurants",
                        total=Decimal("-125.50"),
                        transaction_count=1,
                    ),
                ],
            ),
            CategoryRow(
                category="Transportation",
                total=Decimal("-24.00"),
                percentage=3.5,
                transaction_count=1,
                subcategories=[
                    SubcategoryRow(
                        subcategory="Public Transport",
                        total=Decimal("-24.00"),
                        transaction_count=1,
                    ),
                ],
            ),
        ],
        monthly_data=[
            MonthRow(
                year=2025,
                month=10,
                label="Oct 2025",
                income=Decimal("30000.00"),
                expenses=Decimal("-149.50"),
                net=Decimal("29850.50"),
                transaction_count=3,
            ),
            MonthRow(
                year=2025,
                month=11,
                label="Nov 2025",
                income=Decimal("0"),
                expenses=Decimal("-550.00"),
                net=Decimal("-550.00"),
                transaction_count=2,
            ),
        ],
        source_data=[
            SourceRow(
                source="Danske Bank",
                total_income=Decimal("30000.00"),
                total_expenses=Decimal("-475.50"),
                transaction_count=3,
            ),
            SourceRow(
                source="Nordea",
                total_income=Decimal("0"),
                total_expenses=Decimal("-224.00"),
                transaction_count=2,
            ),
        ],
        period=period,
    )


@pytest.fixture()
def exported_workbook(
    sample_transactions: list[StandardTransaction],
    sample_analytics: AnalyticsResult,
    tmp_path: Path,
) -> openpyxl.Workbook:
    output_file = tmp_path / "test_output.xlsx"
    exporter = ExcelExporter(
        settings=get_settings(),
        analytics_result=sample_analytics,
        output_file=output_file,
    )
    exporter.export(sample_transactions)
    return openpyxl.load_workbook(str(output_file))


class TestExcelExporterBasics:
    def test_export_creates_xlsx_file(
        self,
        sample_transactions: list[StandardTransaction],
        sample_analytics: AnalyticsResult,
        tmp_path: Path,
    ) -> None:
        output_file = tmp_path / "output.xlsx"
        exporter = ExcelExporter(
            settings=get_settings(),
            analytics_result=sample_analytics,
            output_file=output_file,
        )
        exporter.export(sample_transactions)
        assert output_file.exists()

    def test_export_returns_file_path(
        self,
        sample_transactions: list[StandardTransaction],
        sample_analytics: AnalyticsResult,
        tmp_path: Path,
    ) -> None:
        output_file = tmp_path / "output.xlsx"
        exporter = ExcelExporter(
            settings=get_settings(),
            analytics_result=sample_analytics,
            output_file=output_file,
        )
        result = exporter.export(sample_transactions)
        assert result == str(output_file)

    def test_five_sheets_created(self, exported_workbook: openpyxl.Workbook) -> None:
        assert exported_workbook.sheetnames == SHEET_NAMES


class TestTransactionsSheet:
    def test_columns(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Transactions"]
        headers = [ws.cell(1, col).value for col in range(1, 8)]
        assert headers == [
            "Transaction ID",
            "Date",
            "Description",
            "Category",
            "Subcategory",
            "Amount (DKK)",
            "Source",
        ]

    def test_row_count(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Transactions"]
        # 5 transactions + 1 header
        data_rows = [row for row in ws.iter_rows(min_row=2) if row[0].value is not None]
        assert len(data_rows) == 5

    def test_sorted_by_date(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Transactions"]
        dates = [ws.cell(row, 2).value for row in range(2, 7)]
        assert dates == sorted(dates)

    def test_first_transaction_data(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Transactions"]
        # First row (sorted by date) should be Oct 5
        assert ws.cell(2, 2).value == "2025-10-05"
        assert ws.cell(2, 3).value == "Cafe Central"
        assert ws.cell(2, 4).value == "Food & Drinks"


class TestSummarySheet:
    def test_period_label(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Summary"]
        assert ws.cell(1, 1).value == "All Time"

    def test_metric_labels(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Summary"]
        labels = [ws.cell(row, 1).value for row in range(3, 8)]
        assert labels == [
            "Total Transactions",
            "Total Income",
            "Total Expenses",
            "Net",
            "Avg Transaction",
        ]

    def test_metric_values(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Summary"]
        assert ws.cell(3, 2).value == 5  # Total Transactions
        assert ws.cell(4, 2).value == pytest.approx(30000.00)  # Total Income
        assert ws.cell(5, 2).value == pytest.approx(-699.50)  # Total Expenses
        assert ws.cell(6, 2).value == pytest.approx(29300.50)  # Net


class TestCategorySheet:
    def test_headers(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Category Breakdown"]
        headers = [ws.cell(1, col).value for col in range(1, 5)]
        assert headers == ["Category", "Amount (DKK)", "% of Total", "# Transactions"]

    def test_sorted_by_amount_desc(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Category Breakdown"]
        # Car (-350) should come before Food & Drinks (-325.50)
        # sorted by most negative = largest expense first
        assert ws.cell(2, 1).value == "Car"
        assert ws.cell(3, 1).value == "Food & Drinks"

    def test_category_data(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Category Breakdown"]
        assert ws.cell(2, 2).value == pytest.approx(-350.00)
        assert ws.cell(2, 4).value == 1

    def test_has_chart(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Category Breakdown"]
        assert len(ws._charts) >= 1


class TestMonthlySheet:
    def test_headers(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Monthly Trends"]
        headers = [ws.cell(1, col).value for col in range(1, 6)]
        assert headers == ["Month", "Income", "Expenses", "Net", "# Transactions"]

    def test_monthly_data(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Monthly Trends"]
        assert ws.cell(2, 1).value == "Oct 2025"
        assert ws.cell(2, 2).value == pytest.approx(30000.00)
        assert ws.cell(2, 3).value == pytest.approx(-149.50)
        assert ws.cell(3, 1).value == "Nov 2025"

    def test_has_chart(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Monthly Trends"]
        assert len(ws._charts) >= 1


class TestSourceSheet:
    def test_headers(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Source Analysis"]
        headers = [ws.cell(1, col).value for col in range(1, 6)]
        assert headers == ["Source", "Income", "Expenses", "Total", "# Transactions"]

    def test_source_data(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Source Analysis"]
        assert ws.cell(2, 1).value == "Danske Bank"
        assert ws.cell(2, 2).value == pytest.approx(30000.00)
        assert ws.cell(2, 3).value == pytest.approx(-475.50)

    def test_has_chart(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["Source Analysis"]
        assert len(ws._charts) >= 1
